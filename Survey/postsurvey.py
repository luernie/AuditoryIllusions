"""
Post-Experiment Survey
Saves responses to post_survey_responses.xlsx (appends rows if file exists).
Select R (Risset) or S (Shepard) from the dropdown before submitting.
"""

import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import os
from datetime import date

OUTPUT_FILE = "post_survey_responses.xlsx"

OUTPUT_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "post_survey_responses.xlsx"
)

COLUMNS = [
    "Participant ID", "Date", "Experiment",
    "Mental Demand (0-10)",
    "Temporal Demand (0-10)",
    "Performance (0-10)",
    "Effort (0-10)",
    "Frustration (0-10)",
    "Confidence (0-10)",
    "Task Difficulty (0-10)",
    "Ease of Rating — Audio Only (0-10)",
    "Ease of Rating — Haptic Only (0-10)",
    "Ease of Rating — Audio + Haptic (0-10)",
    "Vibration-Only Description",
    "Haptic-Audio Similarity (0-10)",
    "Patterns / Strategies",
    "Confusing or Uncomfortable",
    "Other Comments",
]


def save_response(data: dict):
    if os.path.exists(OUTPUT_FILE):
        wb = openpyxl.load_workbook(OUTPUT_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(COLUMNS)
        for col in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = openpyxl.styles.Font(bold=True)

    row = [data.get(col, "") for col in COLUMNS]
    ws.append(row)
    wb.save(OUTPUT_FILE)


class PostSurveyApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Post-Experiment Survey")
        self.resizable(True, True)
        self.configure(bg="#f5f5f5")
        self.bind("<F11>", lambda e: self.attributes("-fullscreen",
                                                      not self.attributes("-fullscreen")))
        self.bind("<Escape>", lambda e: self.attributes("-fullscreen", False))
        self._build_ui()

    def _section(self, text):
        tk.Label(self.frame, text=text, font=("Arial", 12, "bold"),
                 bg="#dce8f5", anchor="w", padx=8, pady=4,
                 relief="flat").pack(fill="x", pady=(16, 4))

    def _label(self, text, sub=None, wrap=False):
        tk.Label(self.frame, text=text, font=("Arial", 10, "bold"),
                 bg="#f5f5f5", anchor="w", justify="left",
                 wraplength=700 if wrap else 0).pack(fill="x", padx=12, pady=(8, 0))
        if sub:
            tk.Label(self.frame, text=sub, font=("Arial", 9), fg="#555",
                     bg="#f5f5f5", anchor="w").pack(fill="x", padx=12)

    def _slider_row(self, lo_label, hi_label, default=5.0):
        """Canvas slider 0-10, center = 5, ticks at 0, 5, 10."""
        SLIDER_W = 320
        TRACK_Y  = 18
        TICK_H   = 7
        THUMB_R  = 7
        PAD      = THUMB_R + 2

        var = tk.DoubleVar(value=default)

        outer = tk.Frame(self.frame, bg="#f5f5f5")
        outer.pack(padx=20, pady=(4, 0), anchor="w")

        tk.Label(outer, text=lo_label, font=("Arial", 9), fg="#444",
                 bg="#f5f5f5", wraplength=130, justify="right",
                 width=16).pack(side="left")

        canvas = tk.Canvas(outer, width=SLIDER_W, height=44,
                           bg="#f5f5f5", highlightthickness=0)
        canvas.pack(side="left", padx=8)

        tk.Label(outer, text=hi_label, font=("Arial", 9), fg="#444",
                 bg="#f5f5f5", wraplength=130, justify="left",
                 width=16).pack(side="left")

        track_x0 = PAD
        track_x1 = SLIDER_W - PAD
        track_len = track_x1 - track_x0

        def val_to_x(v):
            return track_x0 + v / 10 * track_len

        def x_to_val(x):
            return max(0.0, min(10.0, (x - track_x0) / track_len * 10))

        canvas.create_line(track_x0, TRACK_Y, track_x1, TRACK_Y, fill="#aaa", width=2)

        for i in range(0, 11):
            tx = val_to_x(i)
            canvas.create_line(tx, TRACK_Y, tx, TRACK_Y + TICK_H, fill="#888", width=1)
            if i in (0, 5, 10):
                canvas.create_text(tx, TRACK_Y + TICK_H + 8,
                                   text=str(i), font=("Arial", 8), fill="#444")

        thumb_x = val_to_x(default)
        thumb = canvas.create_oval(thumb_x - THUMB_R, TRACK_Y - THUMB_R,
                                   thumb_x + THUMB_R, TRACK_Y + THUMB_R,
                                   fill="#2a6496", outline="#1e4d72", width=1)

        def redraw(v):
            tx = val_to_x(float(v))
            canvas.coords(thumb, tx - THUMB_R, TRACK_Y - THUMB_R,
                          tx + THUMB_R, TRACK_Y + THUMB_R)

        def jump(event):
            v = x_to_val(event.x)
            var.set(v)
            redraw(v)

        canvas.bind("<Button-1>", jump)
        canvas.bind("<B1-Motion>", jump)
        var.trace_add("write", lambda *_: redraw(var.get()))
        return var

    def _textarea(self, height=4):
        t = tk.Text(self.frame, font=("Arial", 10), width=62, height=height)
        t.pack(padx=20, pady=2, anchor="w")
        return t

    def _build_ui(self):
        outer = tk.Frame(self, bg="#f5f5f5")
        outer.pack(fill="both", expand=True)

        canvas = tk.Canvas(outer, bg="#f5f5f5", highlightthickness=0)
        scrollbar = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        self.frame = tk.Frame(canvas, bg="#f5f5f5", padx=16, pady=16)
        win_id = canvas.create_window((0, 0), window=self.frame, anchor="nw")

        self.frame.bind("<Configure>", lambda _: canvas.configure(
            scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(win_id, width=e.width))
        canvas.bind_all("<MouseWheel>",
                        lambda e: canvas.yview_scroll(-1 * (e.delta // 120), "units"))

        tk.Label(self.frame, text="Post-Experiment Survey",
                 font=("Arial", 16, "bold"), bg="#f5f5f5").pack(pady=(0, 8))

        # Header row — ID, Date, Experiment dropdown
        hf = tk.Frame(self.frame, bg="#f5f5f5")
        hf.pack(anchor="w", pady=4)

        tk.Label(hf, text="Participant ID:", font=("Arial", 10),
                 bg="#f5f5f5").pack(side="left")
        self.pid = tk.Entry(hf, font=("Arial", 10), width=10)
        self.pid.pack(side="left", padx=6)

        tk.Label(hf, text="Date:", font=("Arial", 10),
                 bg="#f5f5f5").pack(side="left", padx=(12, 0))
        self.date_var = tk.Entry(hf, font=("Arial", 10), width=12)
        self.date_var.insert(0, str(date.today()))
        self.date_var.pack(side="left", padx=6)

        tk.Label(hf, text="Experiment:", font=("Arial", 10),
                 bg="#f5f5f5").pack(side="left", padx=(12, 0))
        self.exp_var = tk.StringVar(value="R")
        exp_menu = ttk.Combobox(hf, textvariable=self.exp_var, state="readonly",
                                values=["R", "S"], width=4,
                                font=("Arial", 10))
        exp_menu.pack(side="left", padx=6)

        # Part 1: NASA-TLX
        self._section("PART 1 — NASA TASK LOAD INDEX")
        tk.Label(self.frame,
                 text="Rate each dimension based on your overall experience during the experiment.",
                 font=("Arial", 9), fg="#555", bg="#f5f5f5").pack(padx=12, anchor="w")

        self._label("Mental Demand", "How mentally demanding was the task?")
        self.mental = self._slider_row("Very Low", "Very High")

        self._label("Temporal Demand",
                    "How much time pressure did you feel while completing the task?")
        self.temporal = self._slider_row("Never Rushed", "Always Rushed")

        self._label("Performance",
                    "How successful do you feel you were? (reversed: lower = better performance)")
        self.performance = self._slider_row("Perfect", "Failure")

        self._label("Effort",
                    "How hard did you have to work to achieve your level of performance?")
        self.effort = self._slider_row("Very Low Effort", "Very High Effort")

        self._label("Frustration",
                    "How irritated, stressed, or annoyed did you feel during the task?")
        self.frustration = self._slider_row("Not at All", "Very Much")

        self._label("Confidence",
                    "How confident were you in your judgments during the task?")
        self.confidence = self._slider_row("Not at All Confident", "Extremely Confident")

        self._label("Task Difficulty", "How difficult did you find the task overall?")
        self.difficulty = self._slider_row("Very Easy", "Very Hard")

        # Part 2: Task Experience
        self._section("PART 2 — TASK EXPERIENCE")

        self._label("1. How easy was it to make ratings in each condition?",
                    "0 = very easy, 10 = very hard")
        self._label("    Audio Only", wrap=False)
        self.ease_audio = self._slider_row("Very Easy", "Very Hard")
        self._label("    Haptic Only", wrap=False)
        self.ease_haptic = self._slider_row("Very Easy", "Very Hard")
        self._label("    Audio + Haptic", wrap=False)
        self.ease_both = self._slider_row("Very Easy", "Very Hard")

        self._label("4. When rating the vibration-only condition, describe the types of "
                    "vibrations you encountered. What features did you try to distinguish?",
                    wrap=True)
        self.vib_desc = self._textarea(4)

        self._label("5. Did the vibration sensation feel similar to what you heard in the "
                    "audio condition?",
                    "0 = completely different, 10 = essentially the same")
        self.similarity = self._slider_row("Completely different", "Essentially the same")

        self._label("6. Did you notice any patterns or strategies emerging as you went "
                    "through the experiment?")
        self.strategies = self._textarea(4)

        self._label("7. Was there anything confusing or uncomfortable about the task?")
        self.confusing = self._textarea(4)

        self._label("8. Any other comments about your experience?")
        self.other = self._textarea(4)

        tk.Button(self.frame, text="Submit Survey", font=("Arial", 11, "bold"),
                  bg="#2a6496", fg="white", activebackground="#1e4d72",
                  relief="flat", padx=20, pady=8,
                  command=self._submit).pack(pady=20)

        self.update_idletasks()
        canvas.configure(width=760)
        self.geometry("800x700")

    def _submit(self):
        pid = self.pid.get().strip()
        if not pid:
            messagebox.showwarning("Missing field", "Please enter a Participant ID.")
            return

        exp = self.exp_var.get()
        data = {
            "Participant ID":                         pid,
            "Date":                                   self.date_var.get().strip(),
            "Experiment":                             exp,
            "Mental Demand (0-10)":                   round(self.mental.get(), 2),
            "Temporal Demand (0-10)":                 round(self.temporal.get(), 2),
            "Performance (0-10)":                     round(self.performance.get(), 2),
            "Effort (0-10)":                          round(self.effort.get(), 2),
            "Frustration (0-10)":                     round(self.frustration.get(), 2),
            "Confidence (0-10)":                      round(self.confidence.get(), 2),
            "Task Difficulty (0-10)":                 round(self.difficulty.get(), 2),
            "Ease of Rating — Audio Only (0-10)":     round(self.ease_audio.get(), 2),
            "Ease of Rating — Haptic Only (0-10)":    round(self.ease_haptic.get(), 2),
            "Ease of Rating — Audio + Haptic (0-10)": round(self.ease_both.get(), 2),
            "Vibration-Only Description":             self.vib_desc.get("1.0", "end").strip(),
            "Haptic-Audio Similarity (0-10)":         round(self.similarity.get(), 2),
            "Patterns / Strategies":                  self.strategies.get("1.0", "end").strip(),
            "Confusing or Uncomfortable":             self.confusing.get("1.0", "end").strip(),
            "Other Comments":                         self.other.get("1.0", "end").strip(),
        }

        save_response(data)
        messagebox.showinfo("Saved",
            f"Response for participant '{pid}' (Experiment: {exp}) saved to {OUTPUT_FILE}.")
        self.destroy()


if __name__ == "__main__":
    PostSurveyApp().mainloop()