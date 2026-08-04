"""
Pre-Experiment Survey
Saves responses to pre_survey_responses.xlsx (appends rows if file exists).
"""

import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import os
from datetime import date

OUTPUT_FILE = "pre_survey_responses.xlsx"

OUTPUT_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "pre_survey_responses.xlsx"
)

COLUMNS = [
    "Participant ID", "Date",
    "Age", "Gender", "Dominant Hand",
    "Hearing Impairment", "Hearing Impairment Details",
    "Tactile Impairment", "Tactile Impairment Details",
    "Music Training",
    "Vibration Sensitivity (0-10)",
    "Haptic Experience (0-10)",
]


def save_response(data: dict):
    if os.path.exists(OUTPUT_FILE):
        wb = openpyxl.load_workbook(OUTPUT_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(COLUMNS)
        for col in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = openpyxl.styles.Font(bold=True)

    row = [data.get(col, "") for col in COLUMNS]
    ws.append(row)
    wb.save(OUTPUT_FILE)


class PreSurveyApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Pre-Experiment Survey")
        self.resizable(True, True)
        self.configure(bg="#f5f5f5")
        self.bind("<F11>", lambda e: self.attributes("-fullscreen", not self.attributes("-fullscreen")))
        self.bind("<Escape>", lambda e: self.attributes("-fullscreen", False))
        self._build_ui()

    def _section(self, text):
        tk.Label(self.frame, text=text, font=("Arial", 12, "bold"),
                 bg="#dce8f5", anchor="w", padx=8, pady=4,
                 relief="flat").pack(fill="x", pady=(16, 4))

    def _label(self, text, sub=None, wrap=False):
        tk.Label(self.frame, text=text, font=("Arial", 10, "bold"),
                 bg="#f5f5f5", anchor="w", justify="left",
                 wraplength=700 if wrap else 0).pack(fill="x", padx=12, pady=(8, 0))
        if sub:
            tk.Label(self.frame, text=sub, font=("Arial", 9), fg="#555",
                     bg="#f5f5f5", anchor="w", justify="left",
                     wraplength=700).pack(fill="x", padx=12)

    def _text_entry(self, height=1):
        if height == 1:
            e = tk.Entry(self.frame, font=("Arial", 10), width=60)
            e.pack(padx=20, pady=2, anchor="w")
        else:
            e = tk.Text(self.frame, font=("Arial", 10), width=60, height=height)
            e.pack(padx=20, pady=2, anchor="w")
        return e

    def _pill_row(self, var, options):
        f = tk.Frame(self.frame, bg="#f5f5f5")
        f.pack(padx=20, pady=6, anchor="w")
        buttons = {}

        def select(val):
            var.set(val)
            for v, (btn, _) in buttons.items():
                if v == val:
                    btn.configure(bg="#2a6496", fg="white", relief="groove")
                else:
                    btn.configure(bg="white", fg="#2a6496", relief="groove")

        for txt, val in options:
            btn = tk.Button(f, text=txt, font=("Arial", 10, "bold"), padx=16, pady=6,
                            bg="white", fg="#2a6496", relief="groove",
                            highlightbackground="#2a6496", highlightthickness=2,
                            activebackground="#1e4d72", activeforeground="white",
                            cursor="hand2", borderwidth=2,
                            command=lambda v=val: select(v))
            btn.pack(side="left", padx=5)
            buttons[val] = (btn, txt)
        return f

    def _yes_no_with_detail(self, var, detail_label="Details:"):
        detail_var = tk.StringVar()
        pill_frame = tk.Frame(self.frame, bg="#f5f5f5")
        pill_frame.pack(padx=20, pady=4, anchor="w")
        buttons = {}

        df = tk.Frame(self.frame, bg="#f5f5f5")
        df.pack(padx=20, pady=(0, 4), anchor="w")
        tk.Label(df, text=detail_label, font=("Arial", 9), fg="#555",
                 bg="#f5f5f5").pack(side="left")
        detail_entry = tk.Entry(df, textvariable=detail_var, font=("Arial", 10),
                                width=52, state="disabled")
        detail_entry.pack(side="left", padx=4)

        def select(val):
            var.set(val)
            for v, btn in buttons.items():
                btn.configure(bg="#2a6496" if v == val else "white",
                              fg="white" if v == val else "#2a6496")
            detail_entry.configure(state="normal" if val == "Yes" else "disabled")

        for txt, val in [("No", "No"), ("Yes", "Yes")]:
            btn = tk.Button(pill_frame, text=txt, font=("Arial", 10, "bold"), padx=16, pady=6,
                            bg="white", fg="#2a6496", relief="groove",
                            highlightbackground="#2a6496", highlightthickness=2,
                            activebackground="#1e4d72", activeforeground="white",
                            cursor="hand2", borderwidth=2,
                            command=lambda v=val: select(v))
            btn.pack(side="left", padx=5)
            buttons[val] = btn

        select("No")
        return detail_var

    def _slider_row(self, lo_label, hi_label, default=5.0):
        """Canvas slider 0-10, center = 5, ticks at 0, 5, 10."""
        SLIDER_W = 340
        TRACK_Y  = 18
        TICK_H   = 7
        THUMB_R  = 7
        PAD      = THUMB_R + 2

        var = tk.DoubleVar(value=default)

        outer = tk.Frame(self.frame, bg="#f5f5f5")
        outer.pack(padx=20, pady=(4, 0), anchor="w")

        tk.Label(outer, text=lo_label, font=("Arial", 9), fg="#444",
                 bg="#f5f5f5", wraplength=120, justify="right").pack(side="left")

        canvas = tk.Canvas(outer, width=SLIDER_W, height=44,
                           bg="#f5f5f5", highlightthickness=0)
        canvas.pack(side="left", padx=8)

        tk.Label(outer, text=hi_label, font=("Arial", 9), fg="#444",
                 bg="#f5f5f5", wraplength=120, justify="left").pack(side="left")

        track_x0 = PAD
        track_x1 = SLIDER_W - PAD
        track_len = track_x1 - track_x0

        def val_to_x(v):
            return track_x0 + v / 10 * track_len

        def x_to_val(x):
            return max(0.0, min(10.0, (x - track_x0) / track_len * 10))

        canvas.create_line(track_x0, TRACK_Y, track_x1, TRACK_Y, fill="#aaa", width=2)

        for i in range(0, 11):
            tx = val_to_x(i)
            canvas.create_line(tx, TRACK_Y, tx, TRACK_Y + TICK_H, fill="#888", width=1)
            if i in (0, 5, 10):
                canvas.create_text(tx, TRACK_Y + TICK_H + 8,
                                   text=str(i), font=("Arial", 8), fill="#444")

        thumb_x = val_to_x(default)
        thumb = canvas.create_oval(thumb_x - THUMB_R, TRACK_Y - THUMB_R,
                                   thumb_x + THUMB_R, TRACK_Y + THUMB_R,
                                   fill="#2a6496", outline="#1e4d72", width=1)

        def redraw(v):
            tx = val_to_x(float(v))
            canvas.coords(thumb, tx - THUMB_R, TRACK_Y - THUMB_R,
                          tx + THUMB_R, TRACK_Y + THUMB_R)

        def jump(event):
            v = x_to_val(event.x)
            var.set(v)
            redraw(v)

        canvas.bind("<Button-1>", jump)
        canvas.bind("<B1-Motion>", jump)
        var.trace_add("write", lambda *_: redraw(var.get()))
        return var

    def _build_ui(self):
        outer = tk.Frame(self, bg="#f5f5f5")
        outer.pack(fill="both", expand=True)

        canvas = tk.Canvas(outer, bg="#f5f5f5", highlightthickness=0)
        scrollbar = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        self.frame = tk.Frame(canvas, bg="#f5f5f5", padx=16, pady=16)
        win_id = canvas.create_window((0, 0), window=self.frame, anchor="nw")

        self.frame.bind("<Configure>", lambda _: canvas.configure(
            scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(win_id, width=e.width))
        canvas.bind_all("<MouseWheel>",
                        lambda e: canvas.yview_scroll(-1 * (e.delta // 120), "units"))

        tk.Label(self.frame, text="Pre-Experiment Survey",
                 font=("Arial", 16, "bold"), bg="#f5f5f5").pack(pady=(0, 8))

        hf = tk.Frame(self.frame, bg="#f5f5f5")
        hf.pack(anchor="w", pady=4)
        tk.Label(hf, text="Participant ID:", font=("Arial", 10), bg="#f5f5f5").pack(side="left")
        self.pid = tk.Entry(hf, font=("Arial", 10), width=16)
        self.pid.pack(side="left", padx=6)
        tk.Label(hf, text="Date:", font=("Arial", 10), bg="#f5f5f5").pack(side="left", padx=(16, 0))
        self.date_var = tk.Entry(hf, font=("Arial", 10), width=12)
        self.date_var.insert(0, str(date.today()))
        self.date_var.pack(side="left", padx=6)

        self._section("SECTION 1 — DEMOGRAPHICS")

        self._label("1. Age")
        self.age = self._text_entry()

        self._label("2. Gender", "Free response — or check 'Prefer not to say'.")
        self.gender = self._text_entry()
        self.gender_pnts = tk.BooleanVar()
        pf = tk.Frame(self.frame, bg="#f5f5f5")
        pf.pack(padx=20, anchor="w")
        tk.Checkbutton(pf, text="Prefer not to say", variable=self.gender_pnts,
                       bg="#f5f5f5", font=("Arial", 10)).pack(side="left")

        self._label("3. Dominant Hand")
        self.hand = tk.StringVar(value="Right")
        self._pill_row(self.hand, [("Left", "Left"), ("Right", "Right"),
                                   ("Ambidextrous", "Ambidextrous")])

        self._section("SECTION 2 — SENSORY SCREENING")

        self._label("4. Do you have any hearing impairments?",
                    "Include diagnosed conditions or self-reported difficulty.")
        self.hearing_var = tk.StringVar(value="No")
        self.hearing_detail = self._yes_no_with_detail(self.hearing_var, "Details:")

        self._label("5. Do you have any tactile or somatosensory impairments?",
                    "e.g., reduced sensation in hands or fingers, peripheral neuropathy, etc.")
        self.tactile_var = tk.StringVar(value="No")
        self.tactile_detail = self._yes_no_with_detail(self.tactile_var, "Details:")

        self._section("SECTION 3 — MUSICAL BACKGROUND")

        self._label("6. Do you have any formal music training?",
                    "Describe instruments, ensembles, and approximate years. "
                    "e.g., 'Piano, 5 years classical lessons'. Write N/A if none.", wrap=True)
        self.music = tk.Text(self.frame, font=("Arial", 10), width=60, height=4)
        self.music.pack(padx=20, pady=2, anchor="w")

        self._section("SECTION 4 — PERCEPTUAL SELF-ASSESSMENT")

        self._label("7. How sensitive are you to subtle vibrations in everyday life?",
                    "0 = not at all sensitive, 10 = extremely sensitive")
        self.vib_sensitivity = self._slider_row("Not at all sensitive", "Extremely sensitive")

        self._label("8. How experienced are you with haptic systems or devices?",
                    "0 = no experience at all, 10 = have designed haptic systems")
        self.haptic_exp = self._slider_row("No experience at all", "Have designed haptic systems")

        tk.Button(self.frame, text="Submit Survey", font=("Arial", 11, "bold"),
                  bg="#2a6496", fg="white", activebackground="#1e4d72",
                  relief="flat", padx=20, pady=8, command=self._submit).pack(pady=20)

        self.update_idletasks()
        canvas.configure(width=760)
        self.geometry("800x700")

    def _submit(self):
        pid = self.pid.get().strip()
        if not pid:
            messagebox.showwarning("Missing field", "Please enter a Participant ID.")
            return

        gender_val = "Prefer not to say" if self.gender_pnts.get() else self.gender.get().strip()

        data = {
            "Participant ID":              pid,
            "Date":                        self.date_var.get().strip(),
            "Age":                         self.age.get().strip(),
            "Gender":                      gender_val,
            "Dominant Hand":               self.hand.get(),
            "Hearing Impairment":          self.hearing_var.get(),
            "Hearing Impairment Details":  self.hearing_detail.get().strip(),
            "Tactile Impairment":          self.tactile_var.get(),
            "Tactile Impairment Details":  self.tactile_detail.get().strip(),
            "Music Training":              self.music.get("1.0", "end").strip(),
            "Vibration Sensitivity (0-10)": round(self.vib_sensitivity.get(), 2),
            "Haptic Experience (0-10)":    round(self.haptic_exp.get(), 2),
        }

        save_response(data)
        messagebox.showinfo("Saved",
            f"Response for participant '{pid}' saved to {OUTPUT_FILE}.")
        self.destroy()


if __name__ == "__main__":
    PreSurveyApp().mainloop()